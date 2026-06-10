import streamlit as st
import pandas as pd
import sqlite3
import math
import re
import io

cuposBaseDefault = {'M': 100, 'GA': 50, 'S': 25, 'GCA': 25, 'GL': 20, 'GO': 30}

diccionarioSiglas = {
    'GBIOLO30': 'BIO', 'GBIOQU30': 'BQ', 'GBIOTE30': 'BT',
    'GDFIIE30': 'FIE', 'GELECT30': 'IE', 'GFISIC30': 'FIS',
    'GGEOLO30': 'GEO', 'GINQUI30': 'IQ', 'GMATEM31': 'MAT',
    'GQUIMI30': 'QUIM'
}

def asegura_columnas_originales(conexion):
    cursor = conexion.cursor()
    try:
        cursor.execute('SELECT Grupos_Originales FROM Calculo_Grupos LIMIT 1')
    except sqlite3.OperationalError:
        cursor.execute('ALTER TABLE Calculo_Grupos ADD COLUMN Grupos_Originales INTEGER')
        cursor.execute('ALTER TABLE Calculo_Grupos ADD COLUMN Creditos_Originales REAL')
        
    cursor.execute('''
        UPDATE Calculo_Grupos 
        SET Grupos_Originales = Grupos_Autorizados, 
            Creditos_Originales = Creditos_Autorizados
        WHERE Grupos_Originales IS NULL AND Grupos_Autorizados IS NOT NULL
    ''')
    
    cursor.execute('DELETE FROM Calculo_Grupos WHERE Grupos_Autorizados IS NULL AND Grupos_Originales IS NULL')
    conexion.commit()

def fusiona_diccionarios(dicts):
    res = {}
    for d in dicts:
        for k, v in d.items():
            res[k] = res.get(k, 0) + v
    return res

def distribuye_alumnos_por_grado(dictGrados, numGrupos, tipoClase, cursoMin, cuposActivos):
    totalAlumnos = sum(dictGrados.values())
    if numGrupos <= 0 or totalAlumnos <= 0: return []
    
    cupoBase = cuposActivos.get(tipoClase, 50)
    cupoFisicoAula = cupoBase
    
    if tipoClase == 'S' and str(cursoMin) == '1':
        cupoFisicoAula = 50 
        
    capacidadMinimaNecesaria = math.ceil(totalAlumnos / numGrupos)
    capacidadAula = max(cupoFisicoAula, capacidadMinimaNecesaria)
    
    distribucion = [{'idAula': i, 'capacidadRestante': capacidadAula, 'grados': {}} for i in range(numGrupos)]
    gradosOrdenados = sorted(dictGrados.items(), key = lambda x: x[1], reverse = True)
    
    for grado, cantidad in gradosOrdenados:
        cupoGrado = capacidadAula

        if tipoClase == 'S' and str(cursoMin) == '1':
            if 'GMATEM31' not in grado:
                cupoGrado = max(50, capacidadMinimaNecesaria)
            else:
                cupoGrado = max(cupoBase, capacidadMinimaNecesaria)
        
        tolerancia = 1.15
        if cantidad <= (cupoGrado * tolerancia):
            bloques = [cantidad] 
        else:
            num_trozos = math.ceil(cantidad / cupoGrado)
            base = cantidad // num_trozos
            extra = cantidad % num_trozos
            bloques = [base + 1 if i < extra else base for i in range(num_trozos)]
            
        for alumnosBloque in bloques:
            distribucion.sort(key = lambda x: x['capacidadRestante'], reverse = True)
            distribucion[0]['grados'][grado] = distribucion[0]['grados'].get(grado, 0) + alumnosBloque
            distribucion[0]['capacidadRestante'] -= alumnosBloque
                    
    for gradoEspecifico in dictGrados.keys():
        aulasAsignadas = [aula for aula in distribucion if aula['grados'].get(gradoEspecifico, 0) > 0]
        numeroAulas = len(aulasAsignadas)
        
        if numeroAulas > 1:
            son_todas_puras = all(len(aula['grados']) == 1 for aula in aulasAsignadas)
            
            if son_todas_puras:
                totalAlumnosGrado = sum(aula['grados'][gradoEspecifico] for aula in aulasAsignadas)
                repartoBase = totalAlumnosGrado // numeroAulas
                alumnosSobrantes = totalAlumnosGrado % numeroAulas
                
                for indiceAula, aulaEspecifica in enumerate(aulasAsignadas):
                    asignacionFinal = repartoBase + (1 if indiceAula < alumnosSobrantes else 0)
                    aulaEspecifica['grados'][gradoEspecifico] = asignacionFinal

    distribucion.sort(key = lambda x: x['idAula'])
    return distribucion

def genera_nombre_grupo(tipo, indice, listaGradosEnGrupo):
    prefijo = str(indice)
    siglas = [diccionarioSiglas.get(g, g) for g in sorted(listaGradosEnGrupo)]
    sufijo = '+'.join(siglas)
    return f'{prefijo}{tipo} {sufijo}'

def particiona_agrupacion(dictGrados, mGrupos, cursoMin, cuposActivos):
    if mGrupos > 1:
        distM = distribuye_alumnos_por_grado(dictGrados, mGrupos, 'M', cursoMin, cuposActivos)
        return [d['grados'] for d in distM if sum(d['grados'].values()) > 0]
    return [dictGrados]

def asigna_grupos_a_particiones(particiones, numGruposAutorizados, tipoClase, cursoMin, cuposActivos, esMates = False):
    if not particiones: return []
    if numGruposAutorizados <= 0: return [0] * len(particiones)
    
    necesidades = []
    for part in particiones:
        mat = sum(part.values())
        if mat == 0:
            necesidades.append(0)
            continue
        
        matMat = sum(v for k, v in part.items() if 'GMATEM' in str(k)) if esMates else 0
        matOtros = mat - matMat
        necesidad = calcula_grupos_necesarios(tipoClase, mat, cursoMin, cuposActivos, matMat, matOtros)
        necesidades.append(necesidad)
        
    if sum(necesidades) == numGruposAutorizados:
        return necesidades
        
    asignados = list(necesidades)
    diferencia = numGruposAutorizados - sum(asignados)
    tamanos = [sum(p.values()) for p in particiones]
    
    while diferencia != 0:
        if diferencia > 0:
            idx = max(range(len(tamanos)), key = lambda i: (tamanos[i] / asignados[i]) if asignados[i] > 0 else float('inf'))
            asignados[idx] += 1
            diferencia -= 1
        else:
            idxCandidatos = [i for i, a in enumerate(asignados) if a > 1]
            if idxCandidatos:
                idx = min(idxCandidatos, key = lambda i: tamanos[i] / asignados[i])
                asignados[idx] -= 1
                diferencia += 1
            else:
                break
                
    return asignados

def genera_distribucion_clase(dictGrados, numGrupos, tipoClase, cursoMin, cuposActivos, mGruposAut, esMates):
    if tipoClase == 'M' or mGruposAut <= 1:
        return distribuye_alumnos_por_grado(dictGrados, numGrupos, tipoClase, cursoMin, cuposActivos)
    
    particiones = particiona_agrupacion(dictGrados, mGruposAut, cursoMin, cuposActivos)
    particionesValidas = [p for p in particiones if sum(p.values()) > 0]
    
    if numGrupos < len(particionesValidas):
        dist_fusionada = distribuye_alumnos_por_grado(dictGrados, numGrupos, tipoClase, cursoMin, cuposActivos)
        for i, aula in enumerate(dist_fusionada):
            aula['idAula'] = i
        return dist_fusionada
        
    asigParticiones = asigna_grupos_a_particiones(particiones, numGrupos, tipoClase, cursoMin, cuposActivos, esMates)
    
    distribucionFinal = []
    for part, cant in zip(particiones, asigParticiones):
        if cant > 0:
            distribucionFinal.extend(distribuye_alumnos_por_grado(part, cant, tipoClase, cursoMin, cuposActivos))
            
    for i, aula in enumerate(distribucionFinal):
        aula['idAula'] = i
        
    return distribucionFinal

def crea_id_agrupacion(fila):
    if fila['Compartida'] in ['No', '0', 0]:
        return str(fila['Cod_Plan'])
    grados = re.findall(r'[A-Z]{6}\d{2}', str(fila['Titulaciones_Comparten']))
    grados.sort()
    return '_'.join(grados)

def calcula_grupos_necesarios(tipo, matriculados, curso, cupos, matriculadosMat = 0, matriculadosOtros = 0):
    if pd.isna(matriculados) or matriculados <= 0:
        return 0
    cupo = cupos.get(tipo)
    
    if tipo == 'S' and str(curso) == '1':
        cupoBaseGeneral = 50 
        if matriculadosMat > 0 or matriculadosOtros > 0:
            gruposMat = int(math.ceil(matriculadosMat / cupo)) 
            gruposOtros = int(math.ceil(matriculadosOtros / cupoBaseGeneral))
            return gruposMat + gruposOtros
        else:
            cupo = cupoBaseGeneral
            
    return int(math.ceil(matriculados / cupo))

def calcula_curso_minimo(cursos):
    cursosNums = pd.to_numeric(cursos, errors = 'coerce')
    if cursosNums.isna().all():
        return 'X'
    return int(cursosNums.min())

def calcula_grupos_df(conexion, cupos):
    consulta = '''
    SELECT
        A.Cod_Asignatura, A.Cod_Plan, A.Cod_Dpto, A.Curso, M.Cod_Periodo,
        A.Compartida, A.Titulaciones_Comparten,
        M.Cod_Idioma, M.Matriculados, 
        A.Horas_M, A.Horas_S, A.Horas_GA, A.Horas_GL, A.Horas_GO, A.Horas_GCA
    FROM Matricula M
    LEFT JOIN (
        SELECT * FROM Asignatura_Grado_Dpto 
        GROUP BY Cod_Asignatura, Cod_Plan
    ) A ON M.Cod_Asignatura = A.Cod_Asignatura AND M.Cod_Plan = A.Cod_Plan
    WHERE A.Cod_Asignatura IS NOT NULL
    '''
    dfConsulta = pd.read_sql_query(consulta, conexion)
    
    if dfConsulta.empty:
        return pd.DataFrame()

    dfConsulta['ID_Agrupacion'] = dfConsulta.apply(crea_id_agrupacion, axis = 1)
    dfConsulta = dfConsulta.fillna(0)
    
    dfConsulta['Desglose_Grados'] = dfConsulta.apply(lambda r: {r['Cod_Plan']: r['Matriculados']}, axis = 1)
    
    def es_matematica_especial(fila):
        return str(fila['Cod_Plan']) == 'GMATEM31' and str(fila['Cod_Dpto']).split('.')[0] == '342'
    
    dfConsulta['Es_Mat_Especial'] = dfConsulta.apply(es_matematica_especial, axis = 1)
    dfConsulta['Matriculados_Mat'] = dfConsulta.apply(lambda f: f['Matriculados'] if f['Es_Mat_Especial'] else 0, axis = 1)
    dfConsulta['Matriculados_Otros'] = dfConsulta.apply(lambda f: f['Matriculados'] if not f['Es_Mat_Especial'] else 0, axis = 1)
    
    dfAgrupado = dfConsulta.groupby(
        ['Cod_Asignatura', 'ID_Agrupacion', 'Cod_Periodo', 'Cod_Idioma',
        'Horas_M', 'Horas_S', 'Horas_GA', 'Horas_GL', 'Horas_GO', 'Horas_GCA'] 
    ).agg({
        'Matriculados' : 'sum',
        'Matriculados_Mat' : 'sum',
        'Matriculados_Otros' : 'sum',
        'Curso': calcula_curso_minimo,
        'Desglose_Grados': fusiona_diccionarios, 
        'Es_Mat_Especial': 'any'
    }).reset_index()
    
    filasCalculadas = []
    for _, fila in dfAgrupado.iterrows():
        dictGrados = fila['Desglose_Grados']
        
        mGruposTeorico = 0
        if pd.notna(fila['Horas_M']) and fila['Horas_M'] > 0:
            mGruposTeorico = calcula_grupos_necesarios('M', fila['Matriculados'], fila['Curso'], cupos, fila['Matriculados_Mat'], fila['Matriculados_Otros'])
            
        
        for tipo in cupos.keys():
            if pd.notna(fila[f'Horas_{tipo}']) and fila[f'Horas_{tipo}'] > 0:
                if tipo == 'M':
                    numeroGrupos = mGruposTeorico
                else:
                    numeroGrupos = calcula_grupos_necesarios(tipo, fila['Matriculados'], fila['Curso'], cupos, fila['Matriculados_Mat'], fila['Matriculados_Otros'])

                nuevaFila = {
                    'Cod_Asignatura': fila['Cod_Asignatura'], 
                    'ID_Agrupacion': fila['ID_Agrupacion'], 
                    'Cod_Periodo': fila['Cod_Periodo'], 
                    'Cod_Idioma': fila['Cod_Idioma'], 
                    'Tipo_Clase': tipo, 
                    'Horas': fila[f'Horas_{tipo}'],
                    'Total_Matriculados': fila['Matriculados'],
                    'Desglose_Grados': fila['Desglose_Grados'], 
                    'Grupos_Calculados': numeroGrupos,
                    'Creditos_Calculados': (numeroGrupos * fila[f'Horas_{tipo}']) / 10.0
                }
                filasCalculadas.append(nuevaFila)
                
    dfResultado = pd.DataFrame(filasCalculadas)
    
    if not dfResultado.empty:
        dfExistentes = pd.read_sql_query('SELECT Cod_Asignatura, ID_Agrupacion, Cod_Periodo, Cod_Idioma, Tipo_Clase FROM Calculo_Grupos', conexion)
        if not dfExistentes.empty:
            keys = ['Cod_Asignatura', 'ID_Agrupacion', 'Cod_Periodo', 'Cod_Idioma', 'Tipo_Clase']
            for k in keys:
                dfResultado[k] = dfResultado[k].astype(str).str.strip()
                dfExistentes[k] = dfExistentes[k].astype(str).str.strip()
            
            dfResultado = dfResultado.merge(dfExistentes, on = keys, how = 'inner')
        else:
            dfResultado = pd.DataFrame(columns = dfResultado.columns)
            
    return dfResultado

def actualiza_calculo_grupos(conexion, dfCalculado):
    cursor = conexion.cursor()
    
    for _, fila in dfCalculado.iterrows():
        cursor.execute('''
            UPDATE Calculo_Grupos 
            SET Horas = ?,
                Total_Matriculados = ?,
                Grupos_Calculados = ?,
                Creditos_Calculados = ?
            WHERE Cod_Asignatura = ? AND ID_Agrupacion = ? AND Cod_Periodo = ? AND Cod_Idioma = ? AND Tipo_Clase = ?
        ''', (
            fila['Horas'], fila['Total_Matriculados'], fila['Grupos_Calculados'], fila['Creditos_Calculados'],
            fila['Cod_Asignatura'], fila['ID_Agrupacion'], fila['Cod_Periodo'], fila['Cod_Idioma'], fila['Tipo_Clase']
        ))
        
    conexion.commit()

def guarda_grupos_autorizados(conexion, editedRows, dfOriginal, dfCalculado, cuposActivos):
    cursor = conexion.cursor()
    
    for rowIdxStr, cambios in editedRows.items():
        if 'Grupos_Autorizados' in cambios:
            nuevoValor = cambios['Grupos_Autorizados']
            fila = dfOriginal.iloc[int(rowIdxStr)]
            
            creditosAutorizados = None
            if pd.notna(nuevoValor):
                creditosAutorizados = (nuevoValor * fila['Horas']) / 10.0
            
            cursor.execute('''
                UPDATE Calculo_Grupos 
                SET Grupos_Autorizados = ?, Creditos_Autorizados = ?
                WHERE Cod_Asignatura = ? AND ID_Agrupacion = ? AND Cod_Periodo = ? AND Cod_Idioma = ? AND Tipo_Clase = ?
            ''', (
                nuevoValor, creditosAutorizados, int(fila['Cod_Asignatura']),
                fila['ID_Agrupacion'], fila['Cod_Periodo'], fila['Cod_Idioma'], fila['Tipo_Clase']
            ))
    conexion.commit()

    dfActualizado = pd.read_sql_query('SELECT * FROM Calculo_Grupos', conexion)
    
    for _, fila in dfOriginal.iterrows():
        matchAut = dfActualizado[
            (dfActualizado['Cod_Asignatura'].astype(str) == str(fila['Cod_Asignatura'])) &
            (dfActualizado['ID_Agrupacion'] == fila['ID_Agrupacion']) &
            (dfActualizado['Cod_Periodo'] == fila['Cod_Periodo']) &
            (dfActualizado['Cod_Idioma'] == fila['Cod_Idioma']) &
            (dfActualizado['Tipo_Clase'] == fila['Tipo_Clase'])
        ]
        
        if not matchAut.empty and pd.notna(matchAut.iloc[0]['Grupos_Autorizados']):
            gruposAutorizadosFinales = int(matchAut.iloc[0]['Grupos_Autorizados'])
            
            cursor.execute('''
                DELETE FROM Distribucion_Grupos
                WHERE Cod_Asignatura = ? AND ID_Agrupacion = ? AND Cod_Periodo = ? AND Cod_Idioma = ? AND Tipo_Clase = ?
            ''', (int(fila['Cod_Asignatura']), fila['ID_Agrupacion'], fila['Cod_Periodo'], fila['Cod_Idioma'], fila['Tipo_Clase']))
            
            if gruposAutorizadosFinales > 0:
                matchCalc = dfCalculado[
                    (dfCalculado['Cod_Asignatura'].astype(str) == str(fila['Cod_Asignatura'])) &
                    (dfCalculado['ID_Agrupacion'] == fila['ID_Agrupacion']) &
                    (dfCalculado['Cod_Periodo'] == fila['Cod_Periodo']) &
                    (dfCalculado['Cod_Idioma'] == fila['Cod_Idioma']) &
                    (dfCalculado['Tipo_Clase'] == fila['Tipo_Clase'])
                ]
                
                if not matchCalc.empty:
                    dictGrados = matchCalc.iloc[0]['Desglose_Grados']
                    
                    dfCursos = pd.read_sql_query('SELECT Curso FROM Asignatura_Grado_Dpto WHERE Cod_Asignatura = ?', conexion, params = (int(fila['Cod_Asignatura']),))
                    cursosNums = pd.to_numeric(dfCursos['Curso'], errors = 'coerce')
                    cursoMin = 'X' if cursosNums.isna().all() else int(cursosNums.min())

                    dfDpto = pd.read_sql_query('SELECT Cod_Dpto FROM Asignatura_Grado_Dpto WHERE Cod_Asignatura = ? LIMIT 1', conexion, params = (int(fila['Cod_Asignatura']),))
                    esMates = ('GMATEM31' in fila['ID_Agrupacion']) and (not dfDpto.empty and str(dfDpto.iloc[0]['Cod_Dpto']).split('.')[0] == '342')

                    mGruposAut = 0
                    if fila['Tipo_Clase'] != 'M':
                        mRow = dfActualizado[
                            (dfActualizado['Cod_Asignatura'].astype(str) == str(fila['Cod_Asignatura'])) & 
                            (dfActualizado['ID_Agrupacion'] == fila['ID_Agrupacion']) & 
                            (dfActualizado['Cod_Periodo'] == fila['Cod_Periodo']) & 
                            (dfActualizado['Cod_Idioma'] == fila['Cod_Idioma']) & 
                            (dfActualizado['Tipo_Clase'] == 'M')
                        ]
                        if not mRow.empty and pd.notna(mRow.iloc[0]['Grupos_Autorizados']):
                            mGruposAut = int(mRow.iloc[0]['Grupos_Autorizados'])

                    distribucionFinal = genera_distribucion_clase(dictGrados, gruposAutorizadosFinales, fila['Tipo_Clase'], cursoMin, cuposActivos, mGruposAut, esMates)
                    
                    for i, aula in enumerate(distribucionFinal):
                        indiceReal = i + 1
                        gradosEnEsteAula = list(aula['grados'].keys())
                        alumnosEnEsteAula = sum(aula['grados'].values())
                        nombreOficial = genera_nombre_grupo(fila['Tipo_Clase'], indiceReal, gradosEnEsteAula)
                        
                        cursor.execute('''
                            INSERT INTO Distribucion_Grupos 
                            (Cod_Asignatura, ID_Agrupacion, Cod_Periodo, Cod_Idioma, Tipo_Clase, Num_Grupo, Nombre_Oficial, Alumnos_Asignados)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (
                            int(fila['Cod_Asignatura']), fila['ID_Agrupacion'], fila['Cod_Periodo'], 
                            fila['Cod_Idioma'], fila['Tipo_Clase'], indiceReal, 
                            nombreOficial, alumnosEnEsteAula
                        ))
    conexion.commit()

def recalcula_distribucion_total(conexion, periodo, cuposActivos = None):
    if cuposActivos is None:
        cuposActivos = cuposBaseDefault.copy()
        
    asegura_columnas_originales(conexion)
    
    dfCalculado = calcula_grupos_df(conexion, cuposActivos)
    if dfCalculado.empty:
        return
        
    actualiza_calculo_grupos(conexion, dfCalculado)
    
    dfOriginal = pd.read_sql_query('SELECT * FROM Calculo_Grupos WHERE Cod_Periodo = ?', conexion, params = (periodo,))
    if not dfOriginal.empty:
        guarda_grupos_autorizados(conexion, {}, dfOriginal, dfCalculado, cuposActivos)

def genera_excel_oficial(conexion, dfActual, dfCalculado, filtroPeriodo, filtroAsignatura, filtroAgrupacion, filtroIdioma, cuposActivos):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    libroExcel = openpyxl.Workbook()
    hojaActiva = libroExcel.active
    hojaActiva.title = 'Tabla_Azul'

    azulClaro = '538DD5'
    azulOscuro = '333399'
    rellenoFilaUno = PatternFill(start_color = azulClaro, end_color = azulClaro, fill_type = 'solid')
    rellenoFilaDos = PatternFill(start_color = azulOscuro, end_color = azulOscuro, fill_type = 'solid')
    fuenteBlancaNegrita = Font(name = 'Calibri', size = 11, color = 'FFFFFF', bold = True)

    alineacionCentrada = Alignment(horizontal = 'center', vertical = 'center', wrap_text = True)
    alineacionVertical = Alignment(horizontal = 'center', vertical = 'center', textRotation = 180)

    bordeGrueso = Side(border_style = 'thick', color = '000000')
    bordeFino = Side(border_style = 'thin', color = '000000')

    hojaActiva.row_dimensions[1].height = 80
    hojaActiva.row_dimensions[2].height = 25

    anchosColumna = {
        'A': 18, 'B': 30, 'C': 4, 'D': 8, 'E': 8, 'F': 25, 'G': 12, 'H': 15, 'I': 25, 'J': 10,
        'K': 4.5, 'L': 4.5, 'M': 4.5, 'N': 4.5, 'O': 4.5, 'P': 4.5, 'Q': 4.5, 'R': 4.5
    }
    for indiceColumna in range(19, 39): 
        anchosColumna[get_column_letter(indiceColumna)] = 3.5

    for columna, ancho in anchosColumna.items():
        hojaActiva.column_dimensions[columna].width = ancho

    datosFila1 = {
        'A': 'COSAS ESPECIALES', 'D': 'KURTSOA', 'E': 'ALDAKETA', 'K': 'Matrícula',
        'L': 'Suma', 'M': 'M', 'N': 'GA', 'O': 'S', 'P': 'GO', 'Q': 'GL', 'R': 'GCA'
    }

    for columna, texto in datosFila1.items():
        celda = hojaActiva[f'{columna}1']
        celda.value = texto
        if columna in ['K', 'L']:
            celda.alignment = alineacionVertical
        else:
            celda.alignment = alineacionCentrada

    celdasCombinadas = [
        ('S1:U1', 'GA'), 
        ('V1:Y1', 'S'), 
        ('Z1:AC1', 'GO'),
        ('AD1:AH1', 'GL'), 
        ('AI1:AL1', 'GCA')
    ]
    for rangoCombinacion, texto in celdasCombinadas:
        hojaActiva.merge_cells(rangoCombinacion)
        celdaInicio = hojaActiva[rangoCombinacion.split(':')[0]]
        celdaInicio.value = texto
        celdaInicio.alignment = alineacionCentrada

    datosFila2 = {
        'B': 'Asignaturas', 'F': 'TOTAL', 'G': 'Destino', 'H': 'T',
        'I': 'Por magistrales', 'J': 'GRADO', 'M': 'M', 'N': 'GA', 'O': 'S', 
        'P': 'GO', 'Q': 'GL', 'R': 'GCA'
    }

    for columna, texto in datosFila2.items():
        celda = hojaActiva[f'{columna}2']
        celda.value = texto
        celda.alignment = alineacionCentrada

    bloquesNumericos = {
        'S': (19, 3), 'V': (22, 4), 'Z': (26, 4), 'AD': (30, 5), 'AI': (35, 4)
    }

    for letraInicio, (indiceInicio, cantidad) in bloquesNumericos.items():
        for i in range(cantidad):
            letraColumna = get_column_letter(indiceInicio + i)
            celda = hojaActiva[f'{letraColumna}2']
            celda.value = str(i + 1)
            celda.alignment = alineacionCentrada

    columnasBordeDerechoGrueso = ['R', 'U', 'Y', 'AC', 'AH', 'AL']
    for columna in range(1, 39):
        letraColumna = get_column_letter(columna)
        celdaUno = hojaActiva[f'{letraColumna}1']
        celdaUno.fill = rellenoFilaUno
        celdaUno.font = fuenteBlancaNegrita
        celdaDos = hojaActiva[f'{letraColumna}2']
        celdaDos.fill = rellenoFilaDos
        celdaDos.font = fuenteBlancaNegrita

        bordeUno = Border(left = bordeFino, right = bordeFino, top = bordeFino, bottom = bordeFino)
        bordeDos = Border(left = bordeFino, right = bordeFino, top = bordeFino, bottom = bordeFino)
        if letraColumna in columnasBordeDerechoGrueso:
            bordeUno = Border(left = bordeFino, right = bordeGrueso, top = bordeFino, bottom = bordeFino)
            bordeDos = Border(left = bordeFino, right = bordeGrueso, top = bordeFino, bottom = bordeFino)
        celdaUno.border = bordeUno
        celdaDos.border = bordeDos

    hojaActiva.auto_filter.ref = 'A2:AL3000'

    datosExcel = [] 
    
    dfFiltrado = dfActual.copy()

    if filtroPeriodo != 'Todos': dfFiltrado = dfFiltrado[dfFiltrado['Cod_Periodo'] == filtroPeriodo]
    if filtroAsignatura != 'Todas': dfFiltrado = dfFiltrado[dfFiltrado['Cod_Asignatura'] == filtroAsignatura]
    if filtroAgrupacion != 'Todas': dfFiltrado = dfFiltrado[dfFiltrado['ID_Agrupacion'] == filtroAgrupacion]
    if filtroIdioma != 'Todos': dfFiltrado = dfFiltrado[dfFiltrado['Cod_Idioma'] == filtroIdioma]
        
    dfOrdenado = dfFiltrado.sort_values(by = ['Cod_Asignatura', 'Cod_Idioma', 'ID_Agrupacion'])
    
    try:
        dfDescripcionesAsignaturas = pd.read_sql_query('SELECT Cod_Asignatura, Desc_Asignatura FROM Asignatura', conexion)
        diccionarioDescripciones = dict(zip(dfDescripcionesAsignaturas['Cod_Asignatura'].astype(str), dfDescripcionesAsignaturas['Desc_Asignatura']))
    except:
        diccionarioDescripciones = {}

    asignaturasAgrupadas = dfOrdenado.groupby(['Cod_Asignatura', 'Cod_Periodo', 'Cod_Idioma'], sort = False)
    lista_grupos = list(asignaturasAgrupadas)
    total_grupos = len(lista_grupos)
    
    for indice_grupo, ((codigoAsignatura, codigoPeriodo, codigoIdioma), grupoAsignatura) in enumerate(lista_grupos):
        def suma_segura(valor): return int(valor) if pd.notna(valor) else 0

        totalM = suma_segura(grupoAsignatura[grupoAsignatura['Tipo_Clase'] == 'M']['Grupos_Autorizados'].sum())
        totalGa = suma_segura(grupoAsignatura[grupoAsignatura['Tipo_Clase'] == 'GA']['Grupos_Autorizados'].sum())
        totalS = suma_segura(grupoAsignatura[grupoAsignatura['Tipo_Clase'] == 'S']['Grupos_Autorizados'].sum())
        totalGl = suma_segura(grupoAsignatura[grupoAsignatura['Tipo_Clase'] == 'GL']['Grupos_Autorizados'].sum())
        totalGo = suma_segura(grupoAsignatura[grupoAsignatura['Tipo_Clase'] == 'GO']['Grupos_Autorizados'].sum())
        totalGca = suma_segura(grupoAsignatura[grupoAsignatura['Tipo_Clase'] == 'GCA']['Grupos_Autorizados'].sum())
        
        totalesTextoFinal = f'{totalM}M {totalGa}GA {totalS}S {totalGl}GL {totalGo}GO {totalGca}GCA'
        
        try:
            dfCursos = pd.read_sql_query('SELECT Curso FROM Asignatura_Grado_Dpto WHERE Cod_Asignatura = ?', conexion, params = (codigoAsignatura,))
            cursosNumericos = pd.to_numeric(dfCursos['Curso'], errors = 'coerce')
            cursoValor = 'X' if cursosNumericos.isna().all() else str(int(cursosNumericos.min()))
        except:
            cursoValor = str(codigoPeriodo)
        
        indiceColorParticion = 0 
        agrupacionesAgrupadas = grupoAsignatura.groupby('ID_Agrupacion', sort = False)
        
        for idAgrupacion, grupoAgrupacion in agrupacionesAgrupadas:
            
            cantidadM = suma_segura(grupoAgrupacion[grupoAgrupacion['Tipo_Clase'] == 'M']['Grupos_Autorizados'].sum())
            cantidadGa = suma_segura(grupoAgrupacion[grupoAgrupacion['Tipo_Clase'] == 'GA']['Grupos_Autorizados'].sum())
            cantidadS = suma_segura(grupoAgrupacion[grupoAgrupacion['Tipo_Clase'] == 'S']['Grupos_Autorizados'].sum())
            cantidadGl = suma_segura(grupoAgrupacion[grupoAgrupacion['Tipo_Clase'] == 'GL']['Grupos_Autorizados'].sum())
            cantidadGo = suma_segura(grupoAgrupacion[grupoAgrupacion['Tipo_Clase'] == 'GO']['Grupos_Autorizados'].sum())
            cantidadGca = suma_segura(grupoAgrupacion[grupoAgrupacion['Tipo_Clase'] == 'GCA']['Grupos_Autorizados'].sum())
            
            coincidenciaAgrupacion = dfCalculado[
                (dfCalculado['Cod_Asignatura'].astype(str) == str(codigoAsignatura)) &
                (dfCalculado['ID_Agrupacion'] == idAgrupacion) &
                (dfCalculado['Cod_Idioma'] == codigoIdioma) &
                (dfCalculado['Cod_Periodo'] == codigoPeriodo) 
            ]

            if not coincidenciaAgrupacion.empty:
                diccionarioGrados = coincidenciaAgrupacion.iloc[0]['Desglose_Grados']
            else:
                continue

            try:
                dfDepartamento = pd.read_sql_query('SELECT Cod_Dpto FROM Asignatura_Grado_Dpto WHERE Cod_Asignatura = ? LIMIT 1', conexion, params = (int(codigoAsignatura),))
                esMatematicas = ('GMATEM31' in idAgrupacion) and (not dfDepartamento.empty and str(dfDepartamento.iloc[0]['Cod_Dpto']).split('.')[0] == '342')
            except:
                esMatematicas = False
                
            particiones = particiona_agrupacion(diccionarioGrados, cantidadM, cursoValor, cuposActivos)
            particiones_validas = [p for p in particiones if sum(p.values()) > 0]
            num_validas = len(particiones_validas)
            
            asignacionM = asigna_grupos_a_particiones(particiones, cantidadM, 'M', cursoValor, cuposActivos, esMatematicas)
            
            def obtener_asignacion_o_fusion(cantidad, tipo):
                if 0 < cantidad < num_validas:
                    return [-1] * len(particiones) 
                return asigna_grupos_a_particiones(particiones, cantidad, tipo, cursoValor, cuposActivos, esMatematicas)

            asignacionGa = obtener_asignacion_o_fusion(cantidadGa, 'GA')
            asignacionS = obtener_asignacion_o_fusion(cantidadS, 'S')
            asignacionGl = obtener_asignacion_o_fusion(cantidadGl, 'GL')
            asignacionGo = obtener_asignacion_o_fusion(cantidadGo, 'GO')
            asignacionGca = obtener_asignacion_o_fusion(cantidadGca, 'GCA')

            esPrimeraAgrupacion = True
            
            for indiceParticion, particion in enumerate(particiones):
                
                if codigoIdioma == 'C':
                    tonosColor = ['FF7F50', 'FF6B6B', 'FF69B4'] 
                elif codigoIdioma == 'E':
                    tonosColor = ['98FB98', 'A7F432', '7CCD7C'] 
                elif codigoIdioma == 'I':
                    tonosColor = ['87CEEB', '00FFFF', '99BADD'] 
                else:
                    tonosColor = ['F8F9F9', 'BFC9CA', 'E5E7E9'] 
                    
                colorHexadecimal = tonosColor[indiceColorParticion % 3]
                
                particionM = asignacionM[indiceParticion]
                particionGa = asignacionGa[indiceParticion]
                particionS = asignacionS[indiceParticion]
                particionGl = asignacionGl[indiceParticion]
                particionGo = asignacionGo[indiceParticion]
                particionGca = asignacionGca[indiceParticion]

                listaParticionTexto = []
                if particionM > 0: listaParticionTexto.append(f'{particionM}M')
                if particionGa > 0: listaParticionTexto.append(f'{particionGa}GA')
                elif particionGa == -1: listaParticionTexto.append(f'{cantidadGa}GA(F)')
                if particionS > 0: listaParticionTexto.append(f'{particionS}S')
                elif particionS == -1: listaParticionTexto.append(f'{cantidadS}S(F)')
                if particionGl > 0: listaParticionTexto.append(f'{particionGl}GL')
                elif particionGl == -1: listaParticionTexto.append(f'{cantidadGl}GL(F)')
                if particionGo > 0: listaParticionTexto.append(f'{particionGo}GO')
                elif particionGo == -1: listaParticionTexto.append(f'{cantidadGo}GO(F)')
                if particionGca > 0: listaParticionTexto.append(f'{particionGca}GCA')
                elif particionGca == -1: listaParticionTexto.append(f'{cantidadGca}GCA(F)')
                particionTextoFinal = ' '.join(listaParticionTexto) if listaParticionTexto else ''

                def generar_distribucion_local(part_local, particion_asignada, tipo, cant_total):
                    if particion_asignada == -1:
                        return distribuye_alumnos_por_grado(diccionarioGrados, cant_total, tipo, cursoValor, cuposActivos)
                    return distribuye_alumnos_por_grado(part_local, particion_asignada, tipo, cursoValor, cuposActivos)

                distribucionGa = generar_distribucion_local(particion, particionGa, 'GA', cantidadGa)
                distribucionS = generar_distribucion_local(particion, particionS, 'S', cantidadS)
                distribucionGo = generar_distribucion_local(particion, particionGo, 'GO', cantidadGo)
                distribucionGl = generar_distribucion_local(particion, particionGl, 'GL', cantidadGl)
                distribucionGca = generar_distribucion_local(particion, particionGca, 'GCA', cantidadGca)

                totalMatriculadosParticion = sum(particion.values())
                siglasParticion = '+'.join(sorted([diccionarioSiglas.get(g, g) for g in particion.keys()]))

                esPrimerGradoParticion = True

                for grado, matricula in sorted(particion.items()):
                    def valor_x(distribucionLocal, indiceAAsignar):
                        if 0 <= indiceAAsignar < len(distribucionLocal):
                            return 'X' if distribucionLocal[indiceAAsignar]['grados'].get(grado, 0) > 0 else ''
                        return ''

                    descripcionAsignatura = diccionarioDescripciones.get(str(codigoAsignatura), '')
                    textoAsignatura = f'{codigoAsignatura}-{descripcionAsignatura}' if descripcionAsignatura else str(codigoAsignatura)

                    filaAñadir = [
                        '', 
                        textoAsignatura, 
                        codigoIdioma, 
                        cursoValor, 
                        '', 
                        totalesTextoFinal, 
                        grado if esPrimerGradoParticion else '', 
                        siglasParticion if esPrimerGradoParticion else '', 
                        particionTextoFinal if esPrimerGradoParticion else '', 
                        grado, 
                        matricula, 
                        totalMatriculadosParticion if esPrimerGradoParticion else '', 
                        particionM if esPrimerGradoParticion and particionM > 0 else '', 
                        particionGa if esPrimerGradoParticion and particionGa > 0 else '', 
                        particionS if esPrimerGradoParticion and particionS > 0 else '', 
                        particionGo if esPrimerGradoParticion and particionGo > 0 else '', 
                        particionGl if esPrimerGradoParticion and particionGl > 0 else '', 
                        particionGca if esPrimerGradoParticion and particionGca > 0 else '', 
                        valor_x(distribucionGa, 0), valor_x(distribucionGa, 1), valor_x(distribucionGa, 2), 
                        valor_x(distribucionS, 0), valor_x(distribucionS, 1), valor_x(distribucionS, 2), valor_x(distribucionS, 3), 
                        valor_x(distribucionGo, 0), valor_x(distribucionGo, 1), valor_x(distribucionGo, 2), valor_x(distribucionGo, 3), 
                        valor_x(distribucionGl, 0), valor_x(distribucionGl, 1), valor_x(distribucionGl, 2), valor_x(distribucionGl, 3), valor_x(distribucionGl, 4), 
                        valor_x(distribucionGca, 0), valor_x(distribucionGca, 1), valor_x(distribucionGca, 2), valor_x(distribucionGca, 3) 
                    ]
                    datosExcel.append((filaAñadir, colorHexadecimal, False))
                    
                    esPrimerGradoParticion = False
                    esPrimeraAgrupacion = False
                    
                indiceColorParticion += 1
                
        if indice_grupo < total_grupos - 1:
            siguiente_asignatura = lista_grupos[indice_grupo + 1][0][0]
            if str(codigoAsignatura) != str(siguiente_asignatura):
                filaVacia = [''] * 38
                datosExcel.append((filaVacia, 'FFFFFF', True))

    for valoresFila, colorHexadecimal, esSeparador in datosExcel:
        hojaActiva.append(valoresFila)
        filaActual = hojaActiva.max_row
        
        if not esSeparador:
            colorRelleno = PatternFill(start_color = colorHexadecimal, end_color = colorHexadecimal, fill_type = 'solid')
            for columna in range(1, 39):
                hojaActiva.cell(row = filaActual, column = columna).fill = colorRelleno

    for filaIterada in hojaActiva.iter_rows(min_row = 3, max_row = hojaActiva.max_row, min_col = 1, max_col = 38):
        for celdaIterada in filaIterada:
            celdaIterada.alignment = alineacionCentrada
            if celdaIterada.column_letter in columnasBordeDerechoGrueso:
                celdaIterada.border = Border(left = bordeFino, right = bordeGrueso, top = bordeFino, bottom = bordeFino)
            else:
                celdaIterada.border = Border(left = bordeFino, right = bordeFino, top = bordeFino, bottom = bordeFino)

    return libroExcel

def muestra_pagina_calculo():
    st.markdown('''
        <style>
        .block-container {
            text-align: center;
        }
        h1, h2, h3, h4, h5, h6, p, label {
            text-align: center !important;
            justify-content: center !important;
        }
        div[role="radiogroup"] {
            justify-content: center !important;
            margin: 0 auto;
        }
        [data-testid="stMetric"] {
            display: flex;
            flex-direction: column;
            align-items: center;
        }
        [data-testid="stMetricValue"], [data-testid="stMetricLabel"], [data-testid="stMetricDelta"] {
            text-align: center !important;
            justify-content: center !important;
        }
        .stButton {
            display: flex;
            justify-content: center;
        }
        [data-testid="stDataFrame"] {
            display: flex;
            justify-content: center;
        }
        </style>
    ''', unsafe_allow_html = True)

    st.markdown("<h1 style='text-align: center;'>Cálculo de Grupos</h1>", unsafe_allow_html = True)
    
    if 'bdActual' not in st.session_state or not st.session_state['bdActual']:
        st.warning('Por favor, selecciona una base de datos en la página de Configuración antes de continuar.')
        return
        
    conexion = sqlite3.connect(st.session_state['bdActual'], timeout = 10.0)
    
    try:
        asegura_columnas_originales(conexion)

        if 'cuposActivos' not in st.session_state:
            st.session_state['cuposActivos'] = cuposBaseDefault.copy()
        if 'resetCounter' not in st.session_state:
            st.session_state['resetCounter'] = 0

        st.markdown('### Configuración de Cupos')
        with st.expander('Ver y modificar cupos de alumnos por aula para cada tipo'):
            with st.form('form_cupos'):
                cols = st.columns(len(cuposBaseDefault))
                nuevosCupos = {}
                for i, tipo in enumerate(cuposBaseDefault.keys()):
                    cupoActual = st.session_state['cuposActivos'][tipo]
                    with cols[i]:
                        nuevosCupos[tipo] = st.number_input(
                            f'{tipo}', 
                            min_value = 1, 
                            value = cupoActual, 
                            step = 1, 
                            key = f'cupo_{tipo}_{st.session_state["resetCounter"]}'
                        )
                
                submitted = st.form_submit_button('Confirmar Cambios')
                if submitted:
                    st.session_state['cuposActivos'] = nuevosCupos.copy()
                    st.rerun()

            if st.button('Restablecer cupos por defecto'):
                st.session_state['cuposActivos'] = cuposBaseDefault.copy()
                st.session_state['resetCounter'] += 1
                st.rerun()
                
        st.markdown('---')
        
        with st.spinner('Verificando el estado de los cálculos...'):
            dfCalculado = calcula_grupos_df(conexion, st.session_state['cuposActivos'])
            dfActual = pd.read_sql_query('SELECT * FROM Calculo_Grupos', conexion)
            
            necesitaActualizar = False
            if dfCalculado.empty:
                st.warning('No hay datos suficientes para calcular grupos. Asegúrate de que las asignaturas y matrículas estén cargadas.')
                return

            keys = ['Cod_Asignatura', 'ID_Agrupacion', 'Cod_Periodo', 'Cod_Idioma', 'Tipo_Clase']
            dfCalculado_check = dfCalculado.copy()
            dfActual_check = dfActual.copy()
            
            for k in keys:
                dfCalculado_check[k] = dfCalculado_check[k].astype(str).str.strip()
                dfActual_check[k] = dfActual_check[k].astype(str).str.strip()
                
            mergeDf = dfCalculado_check.merge(dfActual_check, on = keys, how = 'left', indicator = True)
            
            if (mergeDf['_merge'] == 'left_only').any():
                necesitaActualizar = True
            else:
                comunes = mergeDf[mergeDf['_merge'] == 'both']
                if (comunes['Grupos_Calculados_x'] != comunes['Grupos_Calculados_y']).any() or \
                   (comunes['Creditos_Calculados_x'] != comunes['Creditos_Calculados_y']).any() or \
                   (comunes['Horas_x'] != comunes['Horas_y']).any():
                    necesitaActualizar = True

        if necesitaActualizar:
            st.info('Hay grupos sin calcular en la base de datos o los cupos han cambiado.')
            if st.button('Generar / Actualizar Cálculo de Grupos', type = 'primary'):
                with st.spinner('Calculando y guardando grupos en la base de datos...'):
                    actualiza_calculo_grupos(conexion, dfCalculado)
                    st.success('Cálculo generado exitosamente.')
                    st.rerun()
        else:
            st.success('Todos los cálculos de grupos están actualizados y sincronizados con los datos de matrícula y cupos actuales.')

        dfActual = pd.read_sql_query('SELECT * FROM Calculo_Grupos', conexion)
        if not dfActual.empty:
            dfActual['Cod_Asignatura'] = dfActual['Cod_Asignatura'].astype(str)
            
            st.markdown('---')
            st.subheader('Filtros Globales')
            
            if 'realFiltPeriodo' not in st.session_state: st.session_state.realFiltPeriodo = 'Todos'
            if 'realFiltAsig' not in st.session_state: st.session_state.realFiltAsig = 'Todas'
            if 'realFiltAgrup' not in st.session_state: st.session_state.realFiltAgrup = 'Todas'
            if 'realFiltIdioma' not in st.session_state: st.session_state.realFiltIdioma = 'Todos'
            if 'realFiltTipo' not in st.session_state: st.session_state.realFiltTipo = 'Todos'

            currPeriodo = st.session_state.realFiltPeriodo
            currAsig = st.session_state.realFiltAsig
            currAgrup = st.session_state.realFiltAgrup
            currIdioma = st.session_state.realFiltIdioma
            currTipo = st.session_state.realFiltTipo

            def obtiene_opciones_filtradas(columnaExcluida):
                dft = dfActual.copy()
                if columnaExcluida != 'Cod_Periodo' and currPeriodo != 'Todos':
                    dft = dft[dft['Cod_Periodo'] == currPeriodo]
                if columnaExcluida != 'Cod_Asignatura' and currAsig != 'Todas':
                    dft = dft[dft['Cod_Asignatura'] == currAsig]
                if columnaExcluida != 'ID_Agrupacion' and currAgrup != 'Todas':
                    dft = dft[dft['ID_Agrupacion'] == currAgrup]
                if columnaExcluida != 'Cod_Idioma' and currIdioma != 'Todos':
                    dft = dft[dft['Cod_Idioma'] == currIdioma]
                if columnaExcluida != 'Tipo_Clase' and currTipo != 'Todos':
                    dft = dft[dft['Tipo_Clase'] == currTipo]
                return sorted(dft[columnaExcluida].unique().tolist())

            opsPeriodo = ['Todos'] + obtiene_opciones_filtradas('Cod_Periodo')
            opsAsig = ['Todas'] + obtiene_opciones_filtradas('Cod_Asignatura')
            opsAgrup = ['Todas'] + obtiene_opciones_filtradas('ID_Agrupacion')
            opsIdioma = ['Todos'] + obtiene_opciones_filtradas('Cod_Idioma')
            opsTipo = ['Todos'] + obtiene_opciones_filtradas('Tipo_Clase')
            
            if currPeriodo not in opsPeriodo:
                currPeriodo = 'Todos'
                st.session_state.realFiltPeriodo = 'Todos'
            if currAsig not in opsAsig:
                currAsig = 'Todas'
                st.session_state.realFiltAsig = 'Todas'
            if currAgrup not in opsAgrup:
                currAgrup = 'Todas'
                st.session_state.realFiltAgrup = 'Todas'
            if currIdioma not in opsIdioma:
                currIdioma = 'Todos'
                st.session_state.realFiltIdioma = 'Todos'
            if currTipo not in opsTipo:
                currTipo = 'Todos'
                st.session_state.realFiltTipo = 'Todos'

            def actualiza_filtro_periodo(): st.session_state.realFiltPeriodo = st.session_state.uiFiltPeriodo
            def actualiza_filtro_asig(): st.session_state.realFiltAsig = st.session_state.uiFiltAsig
            def actualiza_filtro_agrup(): st.session_state.realFiltAgrup = st.session_state.uiFiltAgrup
            def actualiza_filtro_idioma(): st.session_state.realFiltIdioma = st.session_state.uiFiltIdioma
            def actualiza_filtro_tipo(): st.session_state.realFiltTipo = st.session_state.uiFiltTipo

            col1, col2, col3, col4, col5 = st.columns(5)
            with col1:
                fPeriodo = st.selectbox('Período', opsPeriodo, index = opsPeriodo.index(currPeriodo), key = 'uiFiltPeriodo', on_change = actualiza_filtro_periodo)
            with col2:
                fAsig = st.selectbox('Asignatura', opsAsig, index = opsAsig.index(currAsig), key = 'uiFiltAsig', on_change = actualiza_filtro_asig)
            with col3:
                fAgrup = st.selectbox('Agrupación', opsAgrup, index = opsAgrup.index(currAgrup), key = 'uiFiltAgrup', on_change = actualiza_filtro_agrup)
            with col4:
                fIdioma = st.selectbox('Idioma', opsIdioma, index = opsIdioma.index(currIdioma), key = 'uiFiltIdioma', on_change = actualiza_filtro_idioma)
            with col5:
                fTipo = st.selectbox('Tipo de Clase', opsTipo, index = opsTipo.index(currTipo), key = 'uiFiltTipo', on_change = actualiza_filtro_tipo)

            dfFiltrado = dfActual.copy()
            if currPeriodo != 'Todos':
                dfFiltrado = dfFiltrado[dfFiltrado['Cod_Periodo'] == currPeriodo]
            if currAsig != 'Todas':
                dfFiltrado = dfFiltrado[dfFiltrado['Cod_Asignatura'] == currAsig]
            if currAgrup != 'Todas':
                dfFiltrado = dfFiltrado[dfFiltrado['ID_Agrupacion'] == currAgrup]
            if currIdioma != 'Todos':
                dfFiltrado = dfFiltrado[dfFiltrado['Cod_Idioma'] == currIdioma]
            if currTipo != 'Todos':
                dfFiltrado = dfFiltrado[dfFiltrado['Tipo_Clase'] == currTipo]
                
            fPeriodo = currPeriodo
            fAsig = currAsig
            fAgrup = currAgrup
            fIdioma = currIdioma
            fTipo = currTipo
                
            dfFiltrado['Grupos_Autorizados'] = dfFiltrado['Grupos_Autorizados'].astype('Int64')
            dfFiltrado = dfFiltrado.reset_index(drop = True)
            
            tabCalculo, tabComparativas, tabDistribucion, tabAzul = st.tabs(['Cálculo de Grupos', 'Comparativas', 'Distribución de Grupos', 'Tabla Azul'])
            
            with tabCalculo:
                st.subheader('Resultados y Asignación de Grupos')
                st.markdown('Edita la columna **Grupos_Autorizados** fila por fila. Al finalizar, haz clic en Guardar Cambios.')

                colBtn1, colBtn2 = st.columns(2)
                with colBtn1:
                    if st.button('Restaurar Autorizados', help = 'Restaura los grupos autorizados a los valores adjudicados por la universidad para las filas filtradas.', use_container_width = True):
                        with st.spinner('Restaurando grupos originales...'):
                            cambiosRestaurar = {}
                            for indice, fila in dfFiltrado.iterrows():
                                valOriginal = fila.get('Grupos_Originales', None)
                                if pd.isna(valOriginal):
                                    valOriginal = None
                                cambiosRestaurar[str(indice)] = {'Grupos_Autorizados': valOriginal}
                            
                            guarda_grupos_autorizados(conexion, cambiosRestaurar, dfFiltrado, dfCalculado, st.session_state['cuposActivos'])
                            st.success('Grupos originales restaurados correctamente.')
                        st.rerun()

                with colBtn2:
                    if st.button('Autorizar Calculados', help = 'Aplica los grupos sugeridos por el motor de cálculo a las filas filtradas.', use_container_width = True):
                        with st.spinner('Autorizando grupos y regenerando distribución...'):
                            cambiosAutomaticos = {}
                            for indice, fila in dfFiltrado.iterrows():
                                cambiosAutomaticos[str(indice)] = {'Grupos_Autorizados': fila['Grupos_Calculados']}
                            
                            guarda_grupos_autorizados(conexion, cambiosAutomaticos, dfFiltrado, dfCalculado, st.session_state['cuposActivos'])
                            st.success('Grupos autorizados correctamente.')
                        st.rerun()

                st.data_editor(
                    dfFiltrado,
                    disabled = ['Cod_Asignatura', 'ID_Agrupacion', 'Cod_Periodo', 'Cod_Idioma', 'Tipo_Clase', 'Horas', 
                                'Total_Matriculados', 'Grupos_Calculados', 'Creditos_Calculados', 'Creditos_Autorizados', 'Grupos_Originales', 'Creditos_Originales'],
                    hide_index = True,
                    key = 'data_editor_grupos',
                    use_container_width = True,
                    column_config = {
                        'Grupos_Originales': None,
                        'Creditos_Originales': None
                    }
                )
                
                cambios = st.session_state.get('data_editor_grupos', {}).get('edited_rows', {})
                
                if len(cambios) > 0:
                    st.warning(f'Tienes {len(cambios)} fila(s) con cambios sin guardar.')
                    if st.button('Guardar Cambios Confirmados', type = 'primary'):
                        guarda_grupos_autorizados(conexion, cambios, dfFiltrado, dfCalculado, st.session_state['cuposActivos'])
                        st.success('Cambios guardados correctamente.')
                        st.rerun()

            with tabComparativas:
                tipoComparativa = st.radio(
                    'Selecciona el modo de análisis comparativo:',
                    ['Calculados vs Autorizados (Período Actual)', 'Evolución intra-anual (.1 vs .2)'],
                    horizontal = True
                )
                
                st.markdown('<br>', unsafe_allow_html = True)
                
                if tipoComparativa == 'Calculados vs Autorizados (Período Actual)':
                    st.subheader('Análisis Comparativo: Calculados vs Autorizados')
                    
                    if dfFiltrado.empty:
                        st.info('No hay datos para comparar con los filtros actuales.')
                    else:
                        creditosTeoricos = dfFiltrado['Creditos_Calculados'].sum()
                        creditosReales = dfFiltrado['Creditos_Autorizados'].sum()
                        gruposTeoricos = dfFiltrado['Grupos_Calculados'].sum()
                        gruposReales = dfFiltrado['Grupos_Autorizados'].sum()

                        st.markdown('#### 1. Balance Global')
                        colP1_1, colP1_2 = st.columns(2)
                        
                        with colP1_1:
                            st.metric(label = 'Grupos Calculados (Teoría)', value = f'{gruposTeoricos}')
                        with colP1_2:
                            st.metric(label = 'Créditos Calculados (Teoría)', value = f'{creditosTeoricos:.2f}')

                        colP2_1, colP2_2 = st.columns(2)
                        with colP2_1:
                            if pd.notna(gruposReales) and gruposReales > 0:
                                difGrupos = int(gruposReales - gruposTeoricos)
                                st.metric(label = 'Grupos Autorizados (Real)', value = f'{gruposReales}', delta = f'{difGrupos:+d}' if difGrupos != 0 else None, delta_color = 'inverse')
                            else:
                                st.metric(label = 'Grupos Autorizados (Real)', value = '0')
                        with colP2_2:
                            if pd.notna(creditosReales) and creditosReales > 0:
                                difCreditos = creditosReales - creditosTeoricos
                                st.metric(label = 'Créditos Autorizados (Real)', value = f'{creditosReales:.2f}', delta = f'{difCreditos:+.2f}' if difCreditos != 0 else None, delta_color = 'inverse')
                            else:
                                st.metric(label = 'Créditos Autorizados (Real)', value = '0.00')

                        st.markdown('---')
                        st.markdown('#### 2. Diferencias por Modalidad Docente')

                        dfTipos = dfFiltrado.groupby('Tipo_Clase')[['Grupos_Calculados', 'Grupos_Autorizados', 'Creditos_Calculados', 'Creditos_Autorizados']].sum().reset_index()
                        
                        if not dfTipos.empty:
                            dfTipos['Grupos_Autorizados'] = dfTipos['Grupos_Autorizados'].fillna(0).astype(int)
                            dfTipos['Grupos_Calculados'] = dfTipos['Grupos_Calculados'].fillna(0).astype(int)
                            
                            dfMelted = dfTipos.melt(id_vars = 'Tipo_Clase', value_vars = ['Grupos_Calculados', 'Grupos_Autorizados'], 
                                                    var_name = 'Origen', value_name = 'Cantidad')
                            dfMelted['Origen'] = dfMelted['Origen'].replace({'Grupos_Calculados': 'Calculados (Motor)', 'Grupos_Autorizados': 'Autorizados (Real)'})

                            try:
                                import plotly.express as px
                                fig = px.bar(
                                    dfMelted, x = 'Tipo_Clase', y = 'Cantidad', color = 'Origen', barmode = 'group', text_auto = True,
                                    color_discrete_map = {'Calculados (Motor)': '#636EFA', 'Autorizados (Real)': '#EF553B'},
                                    labels = {'Tipo_Clase': 'Modalidad', 'Cantidad': 'Número de Grupos'}
                                )
                                fig.update_layout(xaxis_title = 'Tipo de Clase', yaxis_title = 'Cantidad de Grupos', legend_title = 'Origen del Dato')
                                st.plotly_chart(fig, use_container_width = True)
                            except ImportError:
                                dfStChart = dfTipos.set_index('Tipo_Clase')[['Grupos_Calculados', 'Grupos_Autorizados']]
                                st.bar_chart(dfStChart)
                            
                            dfTipos['Diferencia_Grupos'] = dfTipos['Grupos_Autorizados'] - dfTipos['Grupos_Calculados']
                            dfTipos['Diferencia_Creditos'] = dfTipos['Creditos_Autorizados'] - dfTipos['Creditos_Calculados']
                            
                            def resaltaDiferencia(col):
                                return ['color: #ff4b4b; font-weight: bold;' if v > 0 else ('color: #21c354; font-weight: bold;' if v < 0 else '') for v in col]
                            
                            def formateaSignoGrupos(v):
                                return f'+{int(v)}' if v > 0 else (f'{int(v)}' if v < 0 else '0')

                            def formateaSignoCreditos(v):
                                return f'+{v:.2f}' if v > 0 else (f'{v:.2f}' if v < 0 else '0.00')
                            
                            st.dataframe(
                                dfTipos.style.apply(resaltaDiferencia, subset = ['Diferencia_Grupos', 'Diferencia_Creditos'])
                                             .format(formateaSignoGrupos, subset = ['Diferencia_Grupos'])
                                             .format(formateaSignoCreditos, subset = ['Diferencia_Creditos'])
                                             .format('{:.2f}', subset = ['Creditos_Calculados', 'Creditos_Autorizados']),
                                use_container_width = True,
                                hide_index = True,
                                column_config = {
                                    'Tipo_Clase': st.column_config.TextColumn('Modalidad Docente', width = 'medium'),
                                    'Grupos_Calculados': st.column_config.NumberColumn('Grupos Calculados', format = '%d'),
                                    'Grupos_Autorizados': st.column_config.NumberColumn('Grupos Autorizados', format = '%d'),
                                    'Diferencia_Grupos': st.column_config.Column('Grupos (Real - Calc)'),
                                    'Creditos_Calculados': st.column_config.NumberColumn('Créditos Calculados', format = '%.2f'),
                                    'Creditos_Autorizados': st.column_config.NumberColumn('Créditos Autorizados', format = '%.2f'),
                                    'Diferencia_Creditos': st.column_config.Column('Créditos (Real - Calc)')
                                }
                            )

                else:
                    st.subheader('Evolución intra-anual: Estimación (.1) vs Final (.2)')
                    
                    if fPeriodo == 'Todos':
                        st.warning('Selecciona un Período específico en los Filtros Globales (ej. 2025/26.2) para identificar el año académico a comparar.')
                    else:
                        anyoBase = str(fPeriodo).split('.')[0]
                        periodo1 = f'{anyoBase}.1'
                        periodo2 = f'{anyoBase}.2'
                        
                        dfComp = dfActual.copy()
                        if fAsig != 'Todas': dfComp = dfComp[dfComp['Cod_Asignatura'] == fAsig]
                        if fAgrup != 'Todas': dfComp = dfComp[dfComp['ID_Agrupacion'] == fAgrup]
                        if fIdioma != 'Todos': dfComp = dfComp[dfComp['Cod_Idioma'] == fIdioma]
                        if fTipo != 'Todos': dfComp = dfComp[dfComp['Tipo_Clase'] == fTipo]
                        
                        dfP1 = dfComp[dfComp['Cod_Periodo'] == periodo1]
                        dfP2 = dfComp[dfComp['Cod_Periodo'] == periodo2]
                        
                        if dfP1.empty or dfP2.empty:
                            st.warning(f'No hay datos suficientes para comparar ambos períodos del año {anyoBase}. Asegúrate de que existen datos autorizados tanto para {periodo1} como para {periodo2}.')
                        else:
                            gruposP1 = dfP1['Grupos_Autorizados'].sum()
                            gruposP2 = dfP2['Grupos_Autorizados'].sum()
                            creditosP1 = dfP1['Creditos_Autorizados'].sum()
                            creditosP2 = dfP2['Creditos_Autorizados'].sum()
                            matP1 = dfP1.drop_duplicates(subset = ['Cod_Asignatura', 'ID_Agrupacion', 'Cod_Idioma', 'Cod_Periodo'])['Total_Matriculados'].sum()
                            matP2 = dfP2.drop_duplicates(subset = ['Cod_Asignatura', 'ID_Agrupacion', 'Cod_Idioma', 'Cod_Periodo'])['Total_Matriculados'].sum()

                            st.markdown('#### 1. Balance Global')
                            colP1_1, colP1_2, colP1_3 = st.columns(3)
                            with colP1_1:
                                st.metric(label = f'Matriculados ({periodo1})', value = int(matP1))
                            with colP1_2:
                                st.metric(label = f'Grupos Autorizados ({periodo1})', value = int(gruposP1) if pd.notna(gruposP1) else 0)
                            with colP1_3:
                                st.metric(label = f'Créditos Autorizados ({periodo1})', value = f'{creditosP1:.2f}')

                            colP2_1, colP2_2, colP2_3 = st.columns(3)
                            with colP2_1:
                                difMat = int(matP2 - matP1)
                                st.metric(label = f'Matriculados ({periodo2})', value = int(matP2), delta = f'{difMat:+d}' if difMat != 0 else None, delta_color = 'normal')
                            with colP2_2:
                                difGruposComp = int(gruposP2 - gruposP1) if pd.notna(gruposP1) and pd.notna(gruposP2) else 0
                                st.metric(label = f'Grupos Autorizados ({periodo2})', value = int(gruposP2) if pd.notna(gruposP2) else 0, delta = f'{difGruposComp:+d}' if difGruposComp != 0 else None, delta_color = 'normal')
                            with colP2_3:
                                difCreditosComp = creditosP2 - creditosP1
                                st.metric(label = f'Créditos Autorizados ({periodo2})', value = f'{creditosP2:.2f}', delta = f'{difCreditosComp:+.2f}' if difCreditosComp != 0 else None, delta_color = 'normal')

                            st.markdown('---')
                            st.markdown('#### 2. Evolución por Modalidad Docente')
                            
                            dfTipos1 = dfP1.groupby('Tipo_Clase')[['Grupos_Autorizados', 'Creditos_Autorizados']].sum().reset_index().rename(columns = {'Grupos_Autorizados': 'Grupos_P1', 'Creditos_Autorizados': 'Creditos_P1'})
                            dfTipos2 = dfP2.groupby('Tipo_Clase')[['Grupos_Autorizados', 'Creditos_Autorizados']].sum().reset_index().rename(columns = {'Grupos_Autorizados': 'Grupos_P2', 'Creditos_Autorizados': 'Creditos_P2'})
                            
                            dfTiposComp = pd.merge(dfTipos1, dfTipos2, on = 'Tipo_Clase', how = 'outer').fillna(0)
                            dfTiposComp['Grupos_P1'] = dfTiposComp['Grupos_P1'].astype(int)
                            dfTiposComp['Grupos_P2'] = dfTiposComp['Grupos_P2'].astype(int)
                            
                            dfMeltedComp = dfTiposComp.melt(id_vars = 'Tipo_Clase', value_vars = ['Grupos_P1', 'Grupos_P2'], var_name = 'Período', value_name = 'Cantidad')
                            dfMeltedComp['Período'] = dfMeltedComp['Período'].map({'Grupos_P1': f'Estimación ({periodo1})', 'Grupos_P2': f'Final ({periodo2})'})
                            
                            try:
                                import plotly.express as px
                                figComp = px.bar(
                                    dfMeltedComp, x = 'Tipo_Clase', y = 'Cantidad', color = 'Período', barmode = 'group', text_auto = True,
                                    color_discrete_map = {f'Estimación ({periodo1})': '#636EFA', f'Final ({periodo2})': '#EF553B'},
                                    labels = {'Tipo_Clase': 'Modalidad', 'Cantidad': 'Número de Grupos'}
                                )
                                figComp.update_layout(xaxis_title = 'Tipo de Clase', yaxis_title = 'Cantidad de Grupos', legend_title = 'Período')
                                st.plotly_chart(figComp, use_container_width = True)
                            except ImportError:
                                dfStChartComp = dfTiposComp.set_index('Tipo_Clase')[['Grupos_P1', 'Grupos_P2']]
                                dfStChartComp.columns = [f'Estimación ({periodo1})', f'Final ({periodo2})']
                                st.bar_chart(dfStChartComp)
                                
                            dfTiposComp['Diferencia_Grupos'] = dfTiposComp['Grupos_P2'] - dfTiposComp['Grupos_P1']
                            dfTiposComp['Diferencia_Creditos'] = dfTiposComp['Creditos_P2'] - dfTiposComp['Creditos_P1']
                            
                            def resaltaDiferenciaComp(col):
                                return ['color: #21c354; font-weight: bold;' if v > 0 else ('color: #ff4b4b; font-weight: bold;' if v < 0 else '') for v in col]
                            
                            def formatSignoGrupos(v):
                                return f'+{int(v)}' if v > 0 else (f'{int(v)}' if v < 0 else '0')

                            def formatSignoCreditos(v):
                                return f'+{v:.2f}' if v > 0 else (f'{v:.2f}' if v < 0 else '0.00')
                                
                            st.dataframe(
                                dfTiposComp.style.apply(resaltaDiferenciaComp, subset = ['Diferencia_Grupos', 'Diferencia_Creditos'])
                                                 .format(formatSignoGrupos, subset = ['Diferencia_Grupos'])
                                                 .format(formatSignoCreditos, subset = ['Diferencia_Creditos'])
                                                 .format('{:.2f}', subset = ['Creditos_P1', 'Creditos_P2']),
                                use_container_width = True,
                                hide_index = True,
                                column_config = {
                                    'Tipo_Clase': st.column_config.TextColumn('Modalidad Docente', width = 'medium'),
                                    'Grupos_P1': st.column_config.NumberColumn(f'Grupos ({periodo1})', format = '%d'),
                                    'Grupos_P2': st.column_config.NumberColumn(f'Grupos ({periodo2})', format = '%d'),
                                    'Diferencia_Grupos': st.column_config.Column('Dif. Grupos (.2 - .1)'),
                                    'Creditos_P1': st.column_config.NumberColumn(f'Créditos ({periodo1})', format = '%.2f'),
                                    'Creditos_P2': st.column_config.NumberColumn(f'Créditos ({periodo2})', format = '%.2f'),
                                    'Diferencia_Creditos': st.column_config.Column('Dif. Créditos (.2 - .1)')
                                }
                            )

            with tabDistribucion:
                st.subheader('Distribución Física de Grupos')
                
                if dfFiltrado.empty:
                    st.info('No hay datos para mostrar.')
                elif dfFiltrado['Grupos_Autorizados'].isna().any():
                    st.warning('Faltan grupos autorizados en la selección filtrada. Por favor, asigna los grupos autorizados en la primera pestaña para continuar.')
                else:
                    def obtener_tabla_azul():
                        df_bruto = pd.read_sql_query('''
                            SELECT 
                                D.Cod_Asignatura, 
                                D.ID_Agrupacion, 
                                D.Cod_Periodo,
                                D.Cod_Idioma, 
                                D.Tipo_Clase, 
                                D.Num_Grupo, 
                                D.Nombre_Oficial, 
                                D.Alumnos_Asignados,
                                C.Total_Matriculados
                            FROM Distribucion_Grupos D
                            JOIN Calculo_Grupos C ON 
                                D.Cod_Asignatura = C.Cod_Asignatura AND
                                D.ID_Agrupacion = C.ID_Agrupacion AND
                                D.Cod_Periodo = C.Cod_Periodo AND
                                D.Cod_Idioma = C.Cod_Idioma AND
                                D.Tipo_Clase = C.Tipo_Clase
                        ''', conexion)

                        df_filt = df_bruto.copy()
                        if not df_filt.empty:
                            df_filt['Cod_Asignatura'] = df_filt['Cod_Asignatura'].astype(str)
                            if fPeriodo != 'Todos': df_filt = df_filt[df_filt['Cod_Periodo'] == fPeriodo]
                            if fAsig != 'Todas': df_filt = df_filt[df_filt['Cod_Asignatura'] == fAsig]
                            if fAgrup != 'Todas': df_filt = df_filt[df_filt['ID_Agrupacion'] == fAgrup]
                            if fIdioma != 'Todos': df_filt = df_filt[df_filt['Cod_Idioma'] == fIdioma]
                            if fTipo != 'Todos': df_filt = df_filt[df_filt['Tipo_Clase'] == fTipo]
                        return df_bruto, df_filt

                    dfTablaAzulBruto, dfTablaAzul = obtener_tabla_azul()

                    if not dfTablaAzul.empty and dfTablaAzul['Alumnos_Asignados'].isna().any():
                        with st.spinner('Generando reparto de alumnos automáticamente...'):
                            guarda_grupos_autorizados(conexion, {}, dfFiltrado, dfCalculado, st.session_state['cuposActivos'])
                            dfTablaAzulBruto, dfTablaAzul = obtener_tabla_azul()

                    if not dfTablaAzulBruto.empty:
                        if not dfTablaAzul.empty:
                            
                            alertas_fusion = []
                            for _, fila_check in dfFiltrado.iterrows():
                                if fila_check['Tipo_Clase'] != 'M' and pd.notna(fila_check['Grupos_Autorizados']) and fila_check['Grupos_Autorizados'] > 0:
                                    m_row = dfFiltrado[
                                        (dfFiltrado['Cod_Asignatura'] == fila_check['Cod_Asignatura']) & 
                                        (dfFiltrado['ID_Agrupacion'] == fila_check['ID_Agrupacion']) & 
                                        (dfFiltrado['Cod_Periodo'] == fila_check['Cod_Periodo']) & 
                                        (dfFiltrado['Cod_Idioma'] == fila_check['Cod_Idioma']) & 
                                        (dfFiltrado['Tipo_Clase'] == 'M')
                                    ]
                                    if not m_row.empty and pd.notna(m_row.iloc[0]['Grupos_Autorizados']):
                                        m_aut = int(m_row.iloc[0]['Grupos_Autorizados'])
                                        if int(fila_check['Grupos_Autorizados']) < m_aut:
                                            alertas_fusion.append(f"Asignatura {fila_check['Cod_Asignatura']} ({fila_check['Cod_Idioma']}): Hay menos grupos de {fila_check['Tipo_Clase']} ({int(fila_check['Grupos_Autorizados'])}) que de Magistral ({m_aut}). El sistema ha mezclado a alumnos de distintas magistrales para asegurar que nadie se quede sin clase.")
                            
                            if alertas_fusion:
                                for alerta in list(dict.fromkeys(alertas_fusion)): 
                                    st.warning(alerta)

                            dfTablaAzul['Ocupacion_Pct'] = dfTablaAzul.apply(
                                lambda x: int((x['Alumnos_Asignados'] / x['Total_Matriculados'] * 100)) if pd.notna(x['Alumnos_Asignados']) and x['Total_Matriculados'] > 0 else 0, 
                                axis = 1
                            )
                            
                            dfTablaAzul['Proporcion_Texto'] = dfTablaAzul.apply(
                                lambda x: f"{int(x['Alumnos_Asignados']) if pd.notna(x['Alumnos_Asignados']) else 0} / {int(x['Total_Matriculados'])}", 
                                axis = 1
                            )

                            st.dataframe(
                                dfTablaAzul,
                                use_container_width = True,
                                hide_index = True,
                                column_config = {
                                    'Cod_Asignatura': 'Asignatura',
                                    'ID_Agrupacion': 'Agrupación',
                                    'Cod_Idioma': 'Idioma',
                                    'Tipo_Clase': 'Tipo',
                                    'Num_Grupo': 'Nº Grupo',
                                    'Nombre_Oficial': st.column_config.TextColumn('Nombre del Grupo (Horarios)', width = 'medium'),
                                    'Total_Matriculados': None, 
                                    'Alumnos_Asignados': None,  
                                    'Proporcion_Texto': st.column_config.TextColumn(
                                        'Alumnos en Aula', 
                                        help = 'Alumnos asignados al aula / Total matriculados en este idioma'
                                    ),
                                    'Ocupacion_Pct': st.column_config.ProgressColumn(
                                        '% Ocupación', 
                                        help = 'Porcentaje de alumnos del idioma asignados a este aula',
                                        format = '%d%%', 
                                        min_value = 0, 
                                        max_value = 100
                                    )
                                }
                            )
                        else:
                            st.info('No hay resultados en la Distribución de Grupos para los filtros seleccionados.')
                    else:
                        st.info('No hay datos para distribuir. Asegúrate de autorizar grupos primero.')


            with tabAzul:
                st.subheader('Tabla Azul')
                
                if fPeriodo == 'Todos':
                    st.warning('Operación bloqueada: Para evitar cruces con datos históricos, debes seleccionar un Período específico en los Filtros Globales antes de generar la Tabla Azul.')
                elif dfFiltrado.empty or dfFiltrado['Grupos_Autorizados'].isna().any() or 'dfTablaAzulBruto' not in locals() or dfTablaAzulBruto.empty or dfTablaAzul.empty:
                    st.info('Autoriza los grupos en la primera pestaña para poder generar la Tabla Azul.')
                else:
                    with st.spinner('Generando vista previa de la Tabla Azul...'):
                        wbOficial = genera_excel_oficial(conexion, dfActual, dfCalculado, fPeriodo, fAsig, fAgrup, fIdioma, st.session_state['cuposActivos'])
                        buffer = io.BytesIO()
                        wbOficial.save(buffer)
                        buffer.seek(0)
                        
                        df_vista_previa = pd.read_excel(buffer, header = None, engine = 'openpyxl')
                        
                        row0 = df_vista_previa.iloc[0].fillna('').astype(str).tolist()
                        row1 = df_vista_previa.iloc[1].fillna('').astype(str).tolist()
                        
                        current_cat = ''
                        final_cols = []
                        seen = {}
                        
                        for c0, c1 in zip(row0, row1):
                            c0 = c0.strip()
                            c1 = c1.strip()
                            
                            if c1.endswith('.0'):
                                c1 = c1[:-2]
                            
                            if c0 != '':
                                current_cat = c0
                                
                            if c1 == '':
                                val = c0 
                            else:
                                if c1.isdigit():
                                    val = f'{current_cat} {c1}'
                                else:
                                    val = c1
                                    
                            if val in seen:
                                seen[val] += 1
                                val = f"{val}{' ' * seen[val]}" 
                            else:
                                seen[val] = 0
                                
                            final_cols.append(val)
                            
                        df_vista_previa.columns = final_cols
                        df_vista_previa = df_vista_previa.iloc[2:].reset_index(drop = True)
                        
                        df_vista_previa = df_vista_previa.fillna('')
                        df_vista_previa = df_vista_previa.astype(str).replace({'nan': '', 'None': '', '<NA>': ''})

                        df_vista_previa = df_vista_previa.replace(r'\.0$', '', regex = True)
                        
                        
                        def aplicar_colores(df):
                            estilos = pd.DataFrame('', index = df.index, columns = df.columns)
                            hoja = wbOficial.active
                            
                            for i in range(len(df)):
                                fila_excel = i + 3
                                celda = hoja.cell(row = fila_excel, column = 2)
                                
                                if celda.fill and celda.fill.start_color and celda.fill.start_color.rgb:
                                    hex_color = str(celda.fill.start_color.rgb)
                                    if hex_color != '00000000' and hex_color != 'FFFFFFFF':
                                        css = f'background-color: #{hex_color[-6:]}; color: black;'
                                        estilos.iloc[i] = css
                            return estilos

                        st.markdown('##### Vista Previa')
                        
                        st.dataframe(
                            df_vista_previa.style.apply(aplicar_colores, axis = None), 
                            use_container_width = True, 
                            hide_index = True
                        )
                        
                        st.markdown('<br>', unsafe_allow_html = True)
                        
                        st.download_button(
                            label = 'Descargar Tabla Azul (.xlsx)',
                            data = buffer.getvalue(),
                            file_name = 'tabla_azul.xlsx',
                            mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                            use_container_width = True
                        )

    finally:
        conexion.close()